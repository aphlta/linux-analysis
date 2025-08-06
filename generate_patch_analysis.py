#!/usr/bin/env python3
import subprocess
import pandas as pd
import re
import argparse
import sys
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import multiprocessing

# 全局缓存
patch_id_cache = {}
cache_lock = threading.Lock()

def run_git_command(cmd):
    """执行git命令并返回输出"""
    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        return result.stdout.strip().split('\n') if result.stdout.strip() else []
    except Exception as e:
        print(f"Error running command: {cmd}")
        print(f"Error: {e}")
        return []

def run_git_command_single(cmd):
    """执行git命令并返回单行输出"""
    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        return result.stdout.strip() if result.returncode == 0 else None
    except Exception as e:
        print(f"Error running command: {cmd}")
        print(f"Error: {e}")
        return None

def validate_branch_exists(branch_name):
    """验证分支是否存在"""
    cmd = f'git rev-parse --verify {branch_name}'
    result = run_git_command_single(cmd)
    return result is not None

def find_merge_base(branch1, branch2):
    """找到两个分支的公共祖先提交"""
    cmd = f'git merge-base {branch1} {branch2}'
    result = run_git_command_single(cmd)
    if result:
        print(f"找到公共祖先提交: {result}")
        return result
    else:
        print(f"警告: 无法找到 {branch1} 和 {branch2} 的公共祖先")
        return None

def clean_text_for_excel(text):
    """清理文本中Excel不支持的字符"""
    if not text:
        return ""

    # 移除或替换Excel不支持的控制字符
    # 保留常见的可打印字符和换行符
    cleaned = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', str(text))

    # 限制文本长度，避免Excel单元格过大
    if len(cleaned) > 32767:  # Excel单元格最大字符数
        cleaned = cleaned[:32760] + "...[截断]"

    return cleaned

def get_patch_id(commit_hash):
    """获取提交的patch-id（带缓存）"""
    with cache_lock:
        if commit_hash in patch_id_cache:
            return patch_id_cache[commit_hash]

    cmd = f'git show {commit_hash} | git patch-id --stable'
    result = run_git_command_single(cmd)
    patch_id = None
    if result:
        patch_id = result.split()[0]  # patch-id是第一个字段

    with cache_lock:
        patch_id_cache[commit_hash] = patch_id

    return patch_id

def build_target_branch_patch_index(target_branch, merge_base=None):
    """构建目标分支的patch-id索引"""
    print(f"正在构建 {target_branch} 分支的patch-id索引...")

    # 如果有公共祖先，只获取公共祖先之后的提交
    if merge_base:
        target_commits_cmd = f'git log {merge_base}..{target_branch} --format=%H'
        print(f"只分析公共祖先 {merge_base[:8]} 之后的提交")
    else:
        target_commits_cmd = f'git log {target_branch} --format=%H'

    target_commits = run_git_command(target_commits_cmd)
    target_commits = [c.strip() for c in target_commits if c.strip()]

    print(f"目标分支有 {len(target_commits)} 个提交需要建立索引")

    # 使用多线程并行获取patch-id
    patch_id_index = set()

    def get_patch_id_for_commit(commit):
        return get_patch_id(commit)

    with ThreadPoolExecutor(max_workers=8) as executor:
        # 提交所有任务
        future_to_commit = {executor.submit(get_patch_id_for_commit, commit): commit
                           for commit in target_commits}

        # 收集结果
        processed = 0
        for future in as_completed(future_to_commit):
            commit = future_to_commit[future]
            try:
                patch_id = future.result()
                if patch_id:
                    patch_id_index.add(patch_id)
                processed += 1

                # 每处理100个提交显示一次进度
                if processed % 100 == 0 or processed == len(target_commits):
                    progress = int(processed / len(target_commits) * 100)
                    print(f"索引构建进度: {progress}% ({processed}/{len(target_commits)})")

            except Exception as e:
                print(f"处理提交 {commit} 时出错: {e}")

    print(f"✅ 索引构建完成，共 {len(patch_id_index)} 个唯一patch-id")
    return patch_id_index

def is_patch_unique_fast(commit_hash, target_patch_index):
    """快速检查补丁是否独有（使用预构建的索引）"""
    source_patch_id = get_patch_id(commit_hash)
    if not source_patch_id:
        print(f"警告: 无法获取提交 {commit_hash} 的patch-id")
        return True  # 如果无法获取patch-id，假设是独有的

    return source_patch_id not in target_patch_index

def get_commit_details(commit_hash):
    """获取提交的详细信息"""
    # 获取完整的提交消息
    full_msg_cmd = f'git log --format="%B" -n 1 {commit_hash}'
    full_message = run_git_command(full_msg_cmd)
    full_message_text = '\n'.join(full_message).strip()

    # 获取修改的文件列表
    files_cmd = f'git diff-tree --no-commit-id --name-only -r {commit_hash}'
    changed_files = run_git_command(files_cmd)
    files_list = ', '.join([f for f in changed_files if f.strip()])

    # 获取文件变更统计
    stats_cmd = f'git diff-tree --no-commit-id --numstat {commit_hash}'
    stats = run_git_command(stats_cmd)

    file_stats = []
    for stat in stats:
        if stat.strip():
            parts = stat.split('\t')
            if len(parts) >= 3:
                added = parts[0] if parts[0] != '-' else '0'
                deleted = parts[1] if parts[1] != '-' else '0'
                filename = parts[2]
                file_stats.append(f"{filename}(+{added}/-{deleted})")

    detailed_files = ', '.join(file_stats)

    return {
        'full_message': clean_text_for_excel(full_message_text),
        'changed_files': clean_text_for_excel(files_list),
        'detailed_files': clean_text_for_excel(detailed_files)
    }

def parse_commit_info(commit_line):
    """解析提交信息行"""
    parts = commit_line.split('|')
    if len(parts) >= 4:
        return {
            'commit_hash': parts[0],
            'author': parts[1],
            'date': parts[2],
            'subject': '|'.join(parts[3:])  # 处理标题中可能包含|的情况
        }
    return None

def categorize_commit(commit_info, changed_files):
    """根据提交信息和修改的文件对提交进行分类"""
    categories = []
    subject = commit_info['subject'].lower()
    files = changed_files.lower()

    # RISC-V相关
    if any(keyword in subject for keyword in ['riscv', 'risc-v']) or 'arch/riscv' in files:
        categories.append('RISC-V')

    # 调度相关 - 更严格的条件
    sched_keywords = [
        'scheduler', 'sched:', 'sched_', 'cfs:', 'cfs_', 'rt:', 'rt_',
        'fair scheduler', 'rt scheduler', 'deadline scheduler',
        'load balancing', 'load balance', 'cpu scheduler',
        'task scheduler', 'process scheduler', 'thread scheduler',
        'runqueue', 'rq_', 'pick_next_task', 'enqueue_task', 'dequeue_task',
        'sched_domain', 'sched_group', 'sched_entity', 'sched_class',
        'wake_up_new_task', 'try_to_wake_up', 'schedule()', 'preempt'
    ]

    # 文件路径检查 - 只有真正的调度相关路径
    sched_paths = [
        'kernel/sched/',
        'include/linux/sched/',
        'include/uapi/linux/sched.h'
    ]

    # 排除明显不是调度的关键词（即使包含sched等词）
    exclude_keywords = [
        # 'usched',  # 用户态调度
        # 'fsched',  # 文件系统调度
        # 'iosched', # IO调度
        # 'elevator', # 电梯调度算法
        # 'deadline', # 除非明确是deadline scheduler
        # 'nohz',    # 除非是调度相关的nohz
        # 'rcu',     # RCU调度
        # 'irq',     # 中断调度
        # 'workqueue', # 工作队列调度
        # 'timer',   # 定时器调度
        # 'hrtimer', # 高精度定时器
        # 'tick',    # 时钟滴答
        # 'clocksource', # 时钟源
        # 'cpufreq', # CPU频率调度
        # 'cpuidle', # CPU空闲
        # 'thermal', # 热调度
        # 'power',   # 电源管理调度
        # 'suspend', # 挂起调度
        # 'hibernate' # 休眠调度
    ]

    # 检查是否为真正的调度相关
    is_scheduler_related = False

    # 首先检查文件路径
    if any(path in files for path in sched_paths):
        is_scheduler_related = True

    # 然后检查关键词，但要排除误报
    elif any(keyword in subject for keyword in sched_keywords):
        # 确保不包含排除的关键词
        if not any(exclude in subject for exclude in exclude_keywords):
            is_scheduler_related = True

    # 特殊情况：如果包含rt但明确是其他子系统的，则排除
    if 'rt' in subject:
        if any(exclude in subject for exclude in ['rtc', 'uart', 'spi', 'i2c', 'usb', 'pci', 'dma', 'gpio']):
            is_scheduler_related = False

    if is_scheduler_related:
        categories.append('调度')

    # 内存管理
    if any(keyword in subject for keyword in ['mm', 'memory', 'page', 'slab', 'kmem']) or any(path in files for path in ['mm/', 'include/linux/mm']):
        categories.append('内存管理')

    # 文件系统
    if any(keyword in subject for keyword in ['fs', 'filesystem', 'ext4', 'btrfs', 'xfs']) or 'fs/' in files:
        categories.append('文件系统')

    # 网络
    if any(keyword in subject for keyword in ['net', 'network', 'tcp', 'udp', 'socket']) or 'net/' in files:
        categories.append('网络')

    # 驱动
    if 'drivers/' in files or any(keyword in subject for keyword in ['driver', 'device']):
        categories.append('驱动')

    # 如果没有匹配到任何分类，归为其他
    if not categories:
        categories.append('其他')

    return categories

def determine_patch_type(subject, full_message):
    """确定补丁类型"""
    subject_lower = subject.lower()
    message_lower = full_message.lower()

    if any(keyword in subject_lower for keyword in ['fix', 'bug', 'error', 'issue']):
        return 'Bug修复'
    elif any(keyword in subject_lower for keyword in ['add', 'new', 'implement', 'introduce']):
        return '新功能'
    elif any(keyword in subject_lower for keyword in ['improve', 'optimize', 'enhance', 'refactor']):
        return '性能优化'
    elif any(keyword in subject_lower for keyword in ['update', 'change', 'modify']):
        return '功能更新'
    else:
        return '其他'

def format_worksheet(worksheet, df):
    """格式化工作表：设置对齐方式、自动换行和自动调整列宽"""
    # 设置对齐方式：左对齐、垂直居中、自动换行
    alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )

    # 设置标题行字体为粗体
    header_font = Font(bold=True)

    # 应用格式到所有单元格
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = alignment
            # 标题行设置粗体
            if cell.row == 1:
                cell.font = header_font

    # 自动调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                # 计算单元格内容的最大长度
                if cell.value:
                    # 处理换行符，取最长的一行
                    lines = str(cell.value).split('\n')
                    cell_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, cell_length)
            except:
                pass

        # 设置列宽，限制最大宽度避免过宽
        adjusted_width = min(max_length + 2, 100)  # 最大100字符宽度
        if adjusted_width < 10:  # 最小10字符宽度
            adjusted_width = 10

        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 设置行高自动调整（对于有换行的内容）
    for row in worksheet.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                max_lines = max(max_lines, lines)

        # 根据行数调整行高（每行大约15个单位）
        if max_lines > 1:
            worksheet.row_dimensions[row[0].row].height = max_lines * 15

def create_formatted_excel(filename, dataframes_dict):
    """创建格式化的Excel文件"""
    wb = Workbook()

    # 删除默认的工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for sheet_name, df in dataframes_dict.items():
        # 创建工作表
        ws = wb.create_sheet(title=sheet_name)

        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # 应用格式
        format_worksheet(ws, df)

    # 保存文件
    wb.save(filename)

def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description='分析两个Git分支之间的独有补丁差异',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""使用示例:
  %(prog)s v6.15.8 openkylin-6.6-next
  %(prog)s --source v6.15.8 --target openkylin-6.6-next
  %(prog)s v6.15.8 openkylin-6.6-next --output my_analysis.xlsx
  %(prog)s v6.15.8 openkylin-6.6-next --no-merge-base
        """
    )

    # 位置参数（可选）
    parser.add_argument('source', nargs='?', help='源分支（要分析其独有补丁的分支）')
    parser.add_argument('target', nargs='?', help='目标分支（用于比较的基准分支）')

    # 命名参数
    parser.add_argument('-s', '--source', dest='source_named', help='源分支（要分析其独有补丁的分支）')
    parser.add_argument('-t', '--target', dest='target_named', help='目标分支（用于比较的基准分支）')
    parser.add_argument('-o', '--output', help='输出文件名（默认: <源分支>-to-<目标分支>-diff.xlsx）')
    parser.add_argument('--no-merge-base', action='store_true', help='不使用公共祖先优化，分析全部差异')
    parser.add_argument('-j', '--jobs', type=int, default=64, help='并行线程数（默认: 64）')
    parser.add_argument('--list-branches', action='store_true', help='列出所有可用的分支')

    args = parser.parse_args()

    # 处理分支参数的优先级：命名参数 > 位置参数
    source_branch = args.source_named or args.source
    target_branch = args.target_named or args.target

    return source_branch, target_branch, args

def list_available_branches():
    """列出所有可用的分支"""
    print("📋 可用的分支列表:")

    # 本地分支
    local_cmd = 'git branch --format="%(refname:short)"'
    local_branches = run_git_command(local_cmd)
    if local_branches:
        print("\n🏠 本地分支:")
        for branch in local_branches:
            if branch.strip():
                print(f"  {branch.strip()}")

    # 远程分支
    remote_cmd = 'git branch -r --format="%(refname:short)"'
    remote_branches = run_git_command(remote_cmd)
    if remote_branches:
        print("\n🌐 远程分支:")
        for branch in remote_branches:
            if branch.strip() and not branch.strip().endswith('/HEAD'):
                print(f"  {branch.strip()}")

    # 标签
    tags_cmd = 'git tag --sort=-version:refname'
    tags = run_git_command(tags_cmd)
    if tags:
        print("\n🏷️  最近的标签 (前10个):")
        for i, tag in enumerate(tags[:10]):
            if tag.strip():
                print(f"  {tag.strip()}")

def get_commit_details_batch(commit_hashes, max_workers=None):
    """批量并行获取多个提交的详细信息"""
    if max_workers is None:
        max_workers = min(len(commit_hashes), multiprocessing.cpu_count())

    results = {}

    def get_details_for_commit(commit_hash):
        return commit_hash, get_commit_details(commit_hash)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_commit = {executor.submit(get_details_for_commit, commit['commit_hash']): commit
                           for commit in commit_hashes}

        for future in as_completed(future_to_commit):
            try:
                commit_hash, details = future.result()
                results[commit_hash] = details
            except Exception as e:
                commit = future_to_commit[future]
                print(f"获取提交 {commit['commit_hash']} 详情时出错: {e}")
                # 提供默认值
                results[commit['commit_hash']] = {
                    'full_message': '',
                    'changed_files': '',
                    'detailed_files': ''
                }

    return results

def analyze_commits_parallel(unique_commits, max_workers=None):
    """并行分析提交（分类和类型确定）"""
    if max_workers is None:
        max_workers = min(len(unique_commits), multiprocessing.cpu_count())

    print(f"🚀 使用 {max_workers} 个线程并行分析提交...")

    # 1. 并行获取所有提交的详细信息
    print("📥 并行获取提交详情...")
    details_dict = get_commit_details_batch(unique_commits, max_workers)

    # 2. 并行进行分类和类型分析
    analysis_data = []
    analysis_lock = threading.Lock()

    def analyze_single_commit(commit):
        commit_hash = commit['commit_hash']
        details = details_dict.get(commit_hash, {
            'full_message': '',
            'changed_files': '',
            'detailed_files': ''
        })

        # 分类
        categories = categorize_commit(commit, details['changed_files'])

        # 确定补丁类型
        patch_type = determine_patch_type(commit['subject'], details['full_message'])

        # 为每个分类创建记录
        local_data = []
        for category in categories:
            local_data.append({
                '提交哈希': commit['commit_hash'],
                '作者': commit['author'],
                '日期': commit['date'],
                '提交标题': commit['subject'],
                '完整提交信息': details['full_message'],
                '修改文件': details['changed_files'],
                '文件变更详情': details['detailed_files'],
                '分类': category,
                '类型': patch_type
            })

        # 线程安全地添加到结果中
        with analysis_lock:
            analysis_data.extend(local_data)

    print("🔍 并行进行分类和类型分析...")
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有分析任务
        futures = [executor.submit(analyze_single_commit, commit) for commit in unique_commits]

        # 等待完成并显示进度
        completed = 0
        for future in as_completed(futures):
            try:
                future.result()
                completed += 1
                if completed % 20 == 0 or completed == len(futures):
                    progress = int(completed / len(futures) * 100)
                    print(f"分析进度: {progress}% ({completed}/{len(futures)})")
            except Exception as e:
                print(f"分析提交时出错: {e}")
                completed += 1

    return analysis_data

def check_unique_commits_parallel(parsed_commits, target_patch_index, max_workers=None):
    """并行检查提交的独有性"""
    if max_workers is None:
        max_workers = min(len(parsed_commits), multiprocessing.cpu_count())

    print(f"🚀 使用 {max_workers} 个线程并行检查独有性...")

    unique_commits = []
    equivalent_count = 0
    results_lock = threading.Lock()

    def check_commit_uniqueness(commit):
        nonlocal equivalent_count
        is_unique = is_patch_unique_fast(commit['commit_hash'], target_patch_index)

        with results_lock:
            if is_unique:
                unique_commits.append(commit)
            else:
                equivalent_count += 1

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有检查任务
        futures = [executor.submit(check_commit_uniqueness, commit) for commit in parsed_commits]

        # 等待完成并显示进度
        completed = 0
        for future in as_completed(futures):
            try:
                future.result()
                completed += 1
                if completed % 50 == 0 or completed == len(futures):
                    progress = int(completed / len(futures) * 100)
                    print(f"独有性检查进度: {progress}% ({completed}/{len(futures)})")
            except Exception as e:
                print(f"检查提交独有性时出错: {e}")
                completed += 1

    return unique_commits, equivalent_count

# 修改主函数中的相关部分
def main():
    # 解析命令行参数
    source_branch, target_branch, args = parse_arguments()

    # 如果请求列出分支
    if args.list_branches:
        list_available_branches()
        return

    # 验证必需的参数
    if not source_branch or not target_branch:
        print("❌ 错误: 必须指定源分支和目标分支")
        print("\n使用方法:")
        print("  python generate_patch_analysis.py <源分支> <目标分支>")
        print("  python generate_patch_analysis.py --source <源分支> --target <目标分支>")
        print("\n查看帮助: python generate_patch_analysis.py --help")
        print("列出分支: python generate_patch_analysis.py --list-branches")
        sys.exit(1)

    # 验证分支是否存在
    if not validate_branch_exists(source_branch):
        print(f"❌ 错误: 源分支 '{source_branch}' 不存在")
        print("\n💡 提示: 使用 --list-branches 查看可用分支")
        sys.exit(1)

    if not validate_branch_exists(target_branch):
        print(f"❌ 错误: 目标分支 '{target_branch}' 不存在")
        print("\n💡 提示: 使用 --list-branches 查看可用分支")
        sys.exit(1)

    # 如果没有指定输出文件名，则动态生成
    if not args.output:
        # 清理分支名称中的特殊字符，避免文件名问题
        safe_source = re.sub(r'[^\w\-_.]', '_', source_branch)
        safe_target = re.sub(r'[^\w\-_.]', '_', target_branch)
        args.output = f"{safe_source}-to-{safe_target}-diff.xlsx"

    print(f"🔍 分析分支差异: {target_branch}..{source_branch}")
    print(f"📁 输出文件: {args.output}")
    print(f"🧵 并行线程: {args.jobs}")

    # 1. 找到公共祖先（除非禁用）
    merge_base = None
    if not args.no_merge_base:
        merge_base = find_merge_base(target_branch, source_branch)

    # 2. 获取源分支相对于公共祖先的提交差异
    if merge_base and not args.no_merge_base:
        commits_cmd = f'git log --pretty=format:"%h|%an|%ad|%s" --date=short --no-merges {merge_base}..{source_branch}'
        print(f"只分析公共祖先 {merge_base[:8]} 之后的提交")
    else:
        commits_cmd = f'git log --pretty=format:"%h|%an|%ad|%s" --date=short --no-merges {target_branch}..{source_branch}'
        print("分析全部差异提交")

    commits = run_git_command(commits_cmd)

    if not commits:
        print("✅ 没有找到提交差异，两个分支内容相同")
        return

    # 解析提交信息
    parsed_commits = []
    for commit in commits:
        if commit.strip():
            commit_info = parse_commit_info(commit)
            if commit_info:
                parsed_commits.append(commit_info)

    total_commits = len(parsed_commits)
    print(f"📊 找到 {total_commits} 个提交需要分析")

    if total_commits == 0:
        print("✅ 没有需要分析的提交")
        return

    # 3. 构建目标分支的patch-id索引
    target_patch_index = build_target_branch_patch_index(target_branch, merge_base)

    # 4. 并行过滤独有补丁（使用预构建的索引）
    print(f"\n🔍 开始并行检查 {total_commits} 个提交的独有性...")

    # 使用并行版本
    unique_commits, equivalent_count = check_unique_commits_parallel(
        parsed_commits, target_patch_index, args.jobs
    )

    print(f"\n📈 过滤结果:")
    print(f"  总提交数: {total_commits}")
    print(f"  等价提交: {equivalent_count}")
    print(f"  独有补丁: {len(unique_commits)}")

    if not unique_commits:
        print("✅ 没有找到独有补丁，所有提交都有等价版本")
        return

    # 5. 并行分类和分析独有补丁
    print(f"\n📝 开始并行分析 {len(unique_commits)} 个独有补丁...")

    # 使用并行版本
    analysis_data = analyze_commits_parallel(unique_commits, args.jobs)

    # 创建DataFrame
    print("📊 创建数据表...")
    df = pd.DataFrame(analysis_data)

    # 按分类统计
    category_stats = df['分类'].value_counts()
    type_stats = df['类型'].value_counts()

    # 准备要写入Excel的数据字典
    excel_data = {
        '独有补丁详情': df,
        '分类统计': pd.DataFrame({
            '分类': category_stats.index,
            '数量': category_stats.values
        }),
        '类型统计': pd.DataFrame({
            '类型': type_stats.index,
            '数量': type_stats.values
        })
    }

    # 添加各分类的专门分析
    for category in category_stats.index:
        category_df = df[df['分类'] == category]
        if not category_df.empty:
            sheet_name = f'{category}独有补丁'
            # Excel工作表名称长度限制
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:28] + '...'
            excel_data[sheet_name] = category_df

    # 创建格式化的Excel文件
    print(f"📄 生成格式化的 {args.output} 文件...")
    try:
        create_formatted_excel(args.output, excel_data)

        print(f"\n🎉 分析完成！")
        print(f"📊 分支对比: {source_branch} vs {target_branch}")
        if merge_base:
            print(f"🔗 公共祖先: {merge_base}")
        print(f"📈 总提交数: {total_commits}")
        print(f"🔄 等价提交: {equivalent_count}")
        print(f"⭐ 独有补丁: {len(unique_commits)}")
        print(f"💾 缓存命中: {len(patch_id_cache)} 个patch-id")
        print(f"🧵 使用线程: {args.jobs}")
        print(f"📁 输出文件: {args.output}")
        print("\n📊 独有补丁分类统计:")
        for category, count in category_stats.items():
            print(f"  {category}: {count} 个补丁")
        print("\n📋 独有补丁类型统计:")
        for patch_type, count in type_stats.items():
            print(f"  {patch_type}: {count} 个补丁")

    except Exception as e:
        print(f"❌ 生成Excel文件时出错: {e}")
        csv_file = args.output.replace('.xlsx', '.csv')
        print(f"尝试保存为CSV格式: {csv_file}")
        df.to_csv(csv_file, index=False, encoding='utf-8-sig')
        print(f"✅ 已保存为CSV文件: {csv_file}")

if __name__ == "__main__":
    main()